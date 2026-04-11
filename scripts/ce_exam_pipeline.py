#!/usr/bin/env python3
"""
国家試験問題抽出・分類・解説生成・画像抽出パイプライン

使用方法:
  # 全工程を実行（問題抽出 → JSON出力）
  python3 scripts/ce_exam_pipeline.py extract --year 2014 --data-dir CE国試data

  # 画像抽出
  python3 scripts/ce_exam_pipeline.py images --year 2014 --data-dir CE国試data

  # xlsx出力（AI分類結果を統合）
  python3 scripts/ce_exam_pipeline.py output --year 2014 --data-dir CE国試data

  # 検証
  python3 scripts/ce_exam_pipeline.py verify --year 2014 --data-dir CE国試data

  # 全工程一括（extract → output）
  python3 scripts/ce_exam_pipeline.py all --year 2014 --data-dir CE国試data
"""

import openpyxl
import json
import re
import os
import sys
import glob
import argparse
from datetime import datetime


# ============================================================
# Step 1: 問題データ抽出
# ============================================================

def extract_questions(filepath):
    """xlsxファイルから問題データを抽出"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Worksheet']

    # 画像位置を取得（0-indexed row）
    image_rows = set()
    for img in ws._images:
        try:
            ref = img.anchor._from
            image_rows.add(ref.row)
        except:
            pass

    questions = []
    rows = list(ws.iter_rows(values_only=False))

    i = 0
    while i < len(rows):
        row_vals = [str(c.value) if c.value else '' for c in rows[i]]
        qid = row_vals[6] if len(row_vals) > 6 else ''

        if qid and re.match(r'^20\d{2}_(?:ce|co|nrs|dh|ort)_', qid):
            category = row_vals[2] if row_vals[2] else ''
            subcategory = row_vals[3] if row_vals[3] else ''

            # 問題文（次の行）
            q_text = ''
            if i + 1 < len(rows):
                q_row = [str(c.value) if c.value else '' for c in rows[i + 1]]
                q_text = q_row[2] if len(q_row) > 2 else ''

            # 画像チェック（問題ブロック内に画像があるか）
            has_image = any(i <= img_row <= i + 10 for img_row in image_rows)

            # 選択肢（最大5つ、画像のみの選択肢にも対応）
            choices = []
            correct_idx = -1
            for j in range(2, 12):
                if i + j >= len(rows):
                    break
                c_row = [str(c.value) if c.value else '' for c in rows[i + j]]
                # 次の問題IDが出たら終了
                if len(c_row) > 6 and c_row[6] and re.match(r'^20\d{2}_(?:ce|co|nrs|dh|ort)_', c_row[6]):
                    break
                choice_num = c_row[2] if len(c_row) > 2 else ''
                if choice_num and choice_num.strip().isdigit():
                    num = int(choice_num.strip())
                    if 1 <= num <= 5:
                        choice_text = c_row[3] if len(c_row) > 3 and c_row[3] else ''
                        if choice_text in ['&#160;', '\xa0', '']:
                            choice_text = f'（図参照：選択肢{num}）'
                            has_image = True
                        is_correct = (len(c_row) > 6 and c_row[6] == '正解')
                        choices.append(choice_text)
                        if is_correct:
                            correct_idx = len(choices) - 1

            correct_letter = chr(ord('A') + correct_idx) if correct_idx >= 0 else ''
            is_multi = any(kw in q_text for kw in ['組合せ', '2つ選べ', '２つ選べ'])
            num_match = re.search(r'(\d+)', qid.split('_')[-1])
            exam_num = int(num_match.group(1)) if num_match else 0

            questions.append({
                'question_id': qid,
                'category': category,
                'subcategory': subcategory,
                'exam_number': exam_num,
                'question_text': q_text,
                'choices': choices,
                'correct_answer': correct_letter,
                'has_image': has_image,
                'is_multi_select': is_multi
            })
        i += 1
    return questions


def cmd_extract(args):
    """問題データ抽出コマンド"""
    data_dir = args.data_dir
    year = args.year

    am_file = pm_file = None
    for f in os.listdir(data_dir):
        if f.endswith('.xlsx') and f'{year}年_午前' in f and 'CE_questions' not in f:
            am_file = os.path.join(data_dir, f)
        elif f.endswith('.xlsx') and f'{year}年_午後' in f and 'CE_questions' not in f:
            pm_file = os.path.join(data_dir, f)

    if not am_file or not pm_file:
        print(f"ERROR: {year}年のxlsxファイルが見つかりません")
        print(f"  午前: {am_file}")
        print(f"  午後: {pm_file}")
        return False

    am_qs = extract_questions(am_file)
    pm_qs = extract_questions(pm_file)
    all_qs = am_qs + pm_qs

    # JSON出力
    output = os.path.join(data_dir, f'{year}_questions_raw.json')
    with open(output, 'w', encoding='utf-8') as f:
        json.dump(all_qs, f, ensure_ascii=False, indent=2)

    # バッチ分割（AI分類用）
    for b in range(6):
        batch = all_qs[b * 30:(b + 1) * 30]
        batch_file = os.path.join(data_dir, f'{year}_batch_{b + 1}.json')
        with open(batch_file, 'w', encoding='utf-8') as f:
            json.dump(batch, f, ensure_ascii=False, indent=2)

    missing = sum(1 for q in all_qs if not q['correct_answer'])
    images = sum(1 for q in all_qs if q['has_image'])
    print(f"{year}年: {len(all_qs)}問抽出 (画像:{images}, 正答なし:{missing})")
    print(f"  出力: {output}")
    print(f"  バッチ: {year}_batch_1.json ~ {year}_batch_6.json")
    return True


# ============================================================
# Step 2: 画像抽出
# ============================================================

def build_question_row_map(filepath):
    """行番号とquestion_idのマッピングを構築"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Worksheet']
    q_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=False)):
        vals = [str(c.value) if c.value else '' for c in row]
        qid = vals[6] if len(vals) > 6 else ''
        if qid and re.match(r'^20\d{2}_(?:ce|co|nrs|dh|ort)_', qid):
            q_rows.append((i, qid))
    return q_rows


def extract_images_for_file(filepath, output_dir):
    """xlsxから画像を抽出しquestion_idに紐づけ"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Worksheet']
    q_rows = build_question_row_map(filepath)
    images = ws._images
    results = []

    for img in images:
        try:
            ref = img.anchor._from
            img_row = ref.row
        except:
            continue

        # どの問題に属するか判定
        assigned_qid = None
        for j, (start_row, qid) in enumerate(q_rows):
            next_start = q_rows[j + 1][0] if j + 1 < len(q_rows) else float('inf')
            if start_row <= img_row < next_start:
                assigned_qid = qid
                break

        if not assigned_qid:
            continue

        # 同一問題の複数画像にサフィックス付与
        existing_for_qid = [r for r in results if r['question_id'] == assigned_qid]
        suffix = f"_{len(existing_for_qid) + 1}" if existing_for_qid else ""

        img_data = img._data()
        ext = 'png'
        if hasattr(img, 'path') and img.path:
            if '.jpg' in img.path or '.jpeg' in img.path:
                ext = 'jpg'
            elif '.gif' in img.path:
                ext = 'gif'

        filename = f"{assigned_qid}{suffix}.{ext}"
        filepath_out = os.path.join(output_dir, filename)

        with open(filepath_out, 'wb') as f:
            f.write(img_data)

        results.append({
            'question_id': assigned_qid,
            'filename': filename,
            'img_row': img_row,
        })
    return results


def cmd_images(args):
    """画像抽出コマンド"""
    data_dir = args.data_dir
    year = args.year
    output_dir = os.path.join(data_dir, 'images')
    os.makedirs(output_dir, exist_ok=True)

    am_file = pm_file = None
    for f in os.listdir(data_dir):
        if f.endswith('.xlsx') and f'{year}年_午前' in f and 'CE_questions' not in f:
            am_file = os.path.join(data_dir, f)
        elif f.endswith('.xlsx') and f'{year}年_午後' in f and 'CE_questions' not in f:
            pm_file = os.path.join(data_dir, f)

    if not am_file or not pm_file:
        print(f"ERROR: {year}年のxlsxファイルが見つかりません")
        return False

    am_results = extract_images_for_file(am_file, output_dir)
    pm_results = extract_images_for_file(pm_file, output_dir)
    all_results = am_results + pm_results

    # 画像マップ作成
    image_map = {}
    for r in all_results:
        qid = r['question_id']
        if qid not in image_map:
            image_map[qid] = []
        image_map[qid].append(r['filename'])

    # マップ保存
    map_file = os.path.join(output_dir, f'{year}_image_map.json')
    with open(map_file, 'w', encoding='utf-8') as f:
        json.dump({
            'year': year,
            'total_images': len(all_results),
            'image_map': image_map,
        }, f, ensure_ascii=False, indent=2)

    print(f"{year}年: {len(all_results)}枚の画像を抽出 ({len(image_map)}問)")
    return True


# ============================================================
# Step 3: xlsx出力（AI分類結果を統合）
# ============================================================

def cmd_output(args):
    """questionsシート形式のxlsx出力"""
    data_dir = args.data_dir
    year = args.year
    dept = args.department

    raw_file = os.path.join(data_dir, f'{year}_questions_raw.json')
    result_file = os.path.join(data_dir, f'{year}_results_merged.json')
    tree_file = os.path.join(data_dir, 'classification_tree.json')
    image_map_file = os.path.join(data_dir, 'images', f'{year}_image_map.json')

    if not os.path.exists(raw_file):
        print(f"ERROR: {raw_file} が見つかりません。先にextractを実行してください。")
        return False

    with open(raw_file, 'r') as f:
        raw = json.load(f)

    # AI結果が存在すれば読み込む
    result_map = {}
    if os.path.exists(result_file):
        with open(result_file, 'r') as f:
            results = json.load(f)
        result_map = {r['question_id']: r for r in results}

    # 分類ツリーで検証
    valid_subtopics = {}
    if os.path.exists(tree_file):
        with open(tree_file, 'r') as f:
            tree = json.load(f)
        for cat, subcats in tree.items():
            for subcat, topics in subcats.items():
                valid_subtopics[(cat, subcat)] = set(topics)

    # subtopic自動修正
    for q in raw:
        qid = q['question_id']
        if qid in result_map:
            r = result_map[qid]
            key = (q['category'], q['subcategory'])
            if key in valid_subtopics and r.get('subtopic') not in valid_subtopics[key]:
                r['subtopic'] = list(valid_subtopics[key])[0]

    # 画像マップ読み込み
    image_map = {}
    if os.path.exists(image_map_file):
        with open(image_map_file, 'r') as f:
            img_data = json.load(f)
        image_map = img_data.get('image_map', {})

    # xlsx作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'questions'
    headers = [
        'question_id', 'department', 'exam_year', 'exam_number',
        'category', 'subcategory', 'subtopic', 'difficulty',
        'question_text', 'choice_a', 'choice_b', 'choice_c', 'choice_d', 'choice_e',
        'correct_answer', 'explanation', 'has_image', 'image_url',
        'is_multi_select', 'source', 'created_at'
    ]
    ws.append(headers)
    now = datetime.now().strftime('%Y-%m-%dT%H:%M:%S.000Z')

    for q in raw:
        r = result_map.get(q['question_id'], {})
        choices = q['choices'] + [''] * (5 - len(q['choices']))
        img_files = image_map.get(q['question_id'], [])
        has_image = len(img_files) > 0
        image_url = ';'.join(img_files) if img_files else ''

        ws.append([
            q['question_id'], dept, year, q['exam_number'],
            q['category'], q['subcategory'],
            r.get('subtopic', ''), r.get('difficulty', 3),
            q['question_text'],
            choices[0], choices[1], choices[2], choices[3], choices[4],
            q['correct_answer'], r.get('explanation', ''),
            str(has_image), image_url,
            str(q['is_multi_select']),
            'past_exam', now
        ])

    # プレフィックスをdepartmentから決定
    prefix_map = {
        'clinical_eng': 'CE', 'orthoptist': 'CO', 'nursing': 'NRS', 'dental_hyg': 'DH'
    }
    prefix = prefix_map.get(dept, dept.upper())
    output_file = os.path.join(data_dir, f'{prefix}_questions_{year}.xlsx')
    wb.save(output_file)
    print(f"{year}年: {ws.max_row - 1}問を {output_file} に出力")
    return True


# ============================================================
# Step 4: 検証
# ============================================================

def cmd_verify(args):
    """出力データの検証"""
    data_dir = args.data_dir
    year = args.year
    dept = getattr(args, 'department', 'clinical_eng')

    prefix_map = {
        'clinical_eng': 'CE', 'orthoptist': 'CO', 'nursing': 'NRS', 'dental_hyg': 'DH'
    }
    prefix = prefix_map.get(dept, dept.upper())
    xlsx_file = os.path.join(data_dir, f'{prefix}_questions_{year}.xlsx')
    if not os.path.exists(xlsx_file):
        print(f"ERROR: {xlsx_file} が見つかりません")
        return False

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    qid_col = headers.index('question_id') + 1
    ca_col = headers.index('correct_answer') + 1
    hi_col = headers.index('has_image') + 1
    iu_col = headers.index('image_url') + 1
    exp_col = headers.index('explanation') + 1
    st_col = headers.index('subtopic') + 1

    issues = []
    total = 0
    for row in range(2, ws.max_row + 1):
        total += 1
        qid = ws.cell(row=row, column=qid_col).value
        ca = ws.cell(row=row, column=ca_col).value
        hi = ws.cell(row=row, column=hi_col).value
        iu = ws.cell(row=row, column=iu_col).value
        exp = ws.cell(row=row, column=exp_col).value
        st = ws.cell(row=row, column=st_col).value

        if not ca:
            issues.append(f"{qid}: 正答なし")
        if not exp:
            issues.append(f"{qid}: 解説なし")
        if not st:
            issues.append(f"{qid}: subtopicなし")
        if hi == 'True' and not iu:
            issues.append(f"{qid}: has_image=True but no image_url")
        if iu:
            for fname in iu.split(';'):
                fpath = os.path.join(data_dir, 'images', fname.strip())
                if not os.path.exists(fpath):
                    issues.append(f"{qid}: 画像ファイル不在 {fname}")

    print(f"{year}年: {total}問を検証")
    if issues:
        print(f"  問題あり: {len(issues)}件")
        for i in issues[:10]:
            print(f"    {i}")
        if len(issues) > 10:
            print(f"    ... 他 {len(issues) - 10}件")
    else:
        print("  問題なし: 全項目OK")
    return len(issues) == 0


# ============================================================
# Step 5: 全年度統合（merge）
# ============================================================

def cmd_merge(args):
    """全年度のCE_questions_{year}.xlsxを1つのxlsxに統合"""
    data_dir = args.data_dir
    dept = args.department
    prefix = args.prefix or dept.upper()

    # 対象ファイルを検出
    pattern = os.path.join(data_dir, f'{prefix}_questions_*.xlsx')
    # CE_questions_2014.xlsx 等にマッチ
    files = sorted(glob.glob(pattern))

    if not files:
        # フォールバック: CE_questions_ でも検索
        pattern2 = os.path.join(data_dir, 'CE_questions_*.xlsx')
        files = sorted(glob.glob(pattern2))

    if not files:
        print(f"ERROR: {pattern} に一致するファイルがありません")
        return False

    # 統合用ワークブック作成
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'questions'

    headers = [
        'question_id', 'department', 'exam_year', 'exam_number',
        'category', 'subcategory', 'subtopic', 'difficulty',
        'question_text', 'choice_a', 'choice_b', 'choice_c', 'choice_d', 'choice_e',
        'correct_answer', 'explanation', 'has_image', 'image_url',
        'is_multi_select', 'source', 'created_at'
    ]
    ws_out.append(headers)

    total = 0
    year_counts = []

    for fpath in files:
        fname = os.path.basename(fpath)
        wb_in = openpyxl.load_workbook(fpath)
        ws_in = wb_in.active

        # ヘッダー行をスキップしてデータ行をコピー
        count = 0
        for row in ws_in.iter_rows(min_row=2, values_only=True):
            ws_out.append(list(row))
            count += 1

        total += count
        year_counts.append((fname, count))
        print(f"  {fname}: {count}問")

    # 統合ファイル保存
    output_file = os.path.join(data_dir, f'{prefix}_questions_all.xlsx')
    wb_out.save(output_file)

    print(f"\n=== 統合完了 ===")
    print(f"  対象ファイル: {len(files)}個")
    print(f"  合計問題数: {total}問")
    print(f"  出力: {output_file}")

    # 年度別サマリーシート追加
    wb_out2 = openpyxl.load_workbook(output_file)
    ws_summary = wb_out2.create_sheet('summary')
    ws_summary.append(['ファイル名', '問題数'])
    for fname, count in year_counts:
        ws_summary.append([fname, count])
    ws_summary.append(['合計', total])
    wb_out2.save(output_file)

    return True


# ============================================================
# メイン
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='国家試験問題処理パイプライン')
    subparsers = parser.add_subparsers(dest='command')

    # 年度指定コマンド
    for name in ['extract', 'images', 'output', 'verify', 'all']:
        sp = subparsers.add_parser(name)
        sp.add_argument('--year', type=int, required=True, help='対象年度')
        sp.add_argument('--data-dir', default='CE国試data', help='データディレクトリ')
        sp.add_argument('--department', default='clinical_eng', help='学科コード')

    # merge コマンド（年度不要）
    sp_merge = subparsers.add_parser('merge', help='全年度のxlsxを1ファイルに統合')
    sp_merge.add_argument('--data-dir', default='CE国試data', help='データディレクトリ')
    sp_merge.add_argument('--department', default='clinical_eng', help='学科コード')
    sp_merge.add_argument('--prefix', default=None, help='ファイルプレフィックス（例: CE）')

    args = parser.parse_args()

    if args.command == 'extract':
        cmd_extract(args)
    elif args.command == 'images':
        cmd_images(args)
    elif args.command == 'output':
        cmd_output(args)
    elif args.command == 'verify':
        cmd_verify(args)
    elif args.command == 'merge':
        cmd_merge(args)
    elif args.command == 'all':
        print(f"=== {args.year}年 全工程実行 ===")
        if cmd_extract(args):
            print("\n--- 画像抽出 ---")
            cmd_images(args)
            print("\n--- xlsx出力 ---")
            cmd_output(args)
            print("\n--- 検証 ---")
            cmd_verify(args)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
