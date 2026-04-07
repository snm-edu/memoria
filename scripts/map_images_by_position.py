"""
xlsxファイル内の画像の配置位置（セル）から問題番号を特定し、
画像ファイルと問題IDを自動マッピングする。

xlsxの構造:
- 各問題は複数行にまたがる（問題文 → 選択肢1 → 選択肢2 → ...）
- 画像は問題文や選択肢のセルに埋め込まれている
- question_idはG列（7列目）の各問題の最初の行に記載
"""
import openpyxl
import zipfile
import os
import json
import re
import glob

data_dir = os.path.join(os.path.dirname(__file__), '..', 'NRS国試data')
out_dir = os.path.join(os.path.dirname(__file__), '..', 'pwa-frontend', 'public', 'images', 'questions')
output_path = os.path.join(os.path.dirname(__file__), 'output', 'image_map.json')

# 既存のマッピングを読み込み
if os.path.exists(output_path):
    with open(output_path, 'r') as f:
        image_map = json.load(f)
else:
    image_map = {}

print(f'既存マッピング: {len(image_map)}件')

xlsx_files = sorted(glob.glob(os.path.join(data_dir, '*.xlsx')))

for fpath in xlsx_files:
    fname = os.path.basename(fpath)
    # 模擬試験はスキップ
    if '模擬' in fname:
        continue

    try:
        wb = openpyxl.load_workbook(fpath)
        sheet = wb.active
        images = sheet._images

        if not images:
            continue

        # 年度とam/pmを判定
        year_match = re.search(r'(\d{4})年', fname)
        if not year_match:
            continue
        year = year_match.group(1)
        year_short = year[2:]  # "14", "15", etc.

        is_pm = '午後' in fname
        period = 'pm' if is_pm else 'am'

        # 問題IDとその開始行を特定
        # G列（7列目）にquestion_idがある
        question_rows = []  # [(row_number, question_id), ...]
        for row in range(1, sheet.max_row + 1):
            val = sheet.cell(row=row, column=7).value
            if val and isinstance(val, str) and ('nrs_' in val or 'NRS' in val.upper()):
                question_rows.append((row, str(val).strip()))

        if not question_rows:
            # G列にIDがない場合、問題番号から推測
            # C列に問題番号（1, 2, 3...）が入っている場合がある
            for row in range(1, sheet.max_row + 1):
                c3 = sheet.cell(row=row, column=3).value
                if c3 and isinstance(c3, str) and '。' in c3 and len(c3) > 10:
                    # 問題文っぽい行を検出
                    # 前の行の選択肢番号から問題番号を推測
                    pass

        # 画像のアンカー位置から問題を特定
        # openpyxlのTwoCellAnchorまたはOneCellAnchor
        from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, OneCellAnchor

        # zipからメディアファイル名のリストを取得
        with zipfile.ZipFile(fpath, 'r') as z:
            media_files = sorted([f for f in z.namelist() if f.startswith('xl/media/')])

        mapped_count = 0
        for idx, img in enumerate(images):
            # 画像のアンカー行を取得
            anchor_row = None
            try:
                if hasattr(img.anchor, '_from'):
                    anchor_row = img.anchor._from.row + 1  # 0-indexed → 1-indexed
                elif hasattr(img.anchor, 'pos'):
                    # AbsoluteAnchorの場合は位置から推測
                    pass
            except:
                pass

            if anchor_row is None:
                continue

            # この行がどの問題に属するか特定
            matched_qid = None
            for i, (qrow, qid) in enumerate(question_rows):
                next_qrow = question_rows[i + 1][0] if i + 1 < len(question_rows) else sheet.max_row + 1
                if qrow <= anchor_row < next_qrow:
                    matched_qid = qid
                    break

            if not matched_qid:
                # question_rowsが空の場合、行番号から問題番号を推測
                # 一般的に1問あたり約6行（問題文+選択肢4-5個）
                estimated_q_num = (anchor_row - 1) // 6 + 1
                matched_qid = f'{year}_nrs_{year_short}_{period}{str(estimated_q_num).zfill(3)}'

            # 対応するメディアファイル
            if idx < len(media_files):
                img_filename = os.path.basename(media_files[idx])

                # 既にnrs_で始まるファイル名ならスキップ（既にマッピング済み）
                if img_filename.startswith('nrs_') and matched_qid in image_map:
                    continue

                # マッピングに追加（まだマッピングされていない場合）
                if matched_qid not in image_map:
                    image_map[matched_qid] = img_filename
                    mapped_count += 1
                elif image_map[matched_qid] != img_filename:
                    # 既に別の画像がマッピングされている場合、追加画像として保存
                    extra_key = f'{matched_qid}_img{idx+1}'
                    image_map[extra_key] = img_filename
                    mapped_count += 1

        if mapped_count > 0:
            print(f'{fname}: 新規{mapped_count}件マッピング')

        wb.close()

    except Exception as e:
        print(f'エラー: {fname}: {e}')

# 保存
with open(output_path, 'w') as f:
    json.dump(image_map, f, indent=2, ensure_ascii=False)

print(f'\n=== 最終結果 ===')
print(f'総マッピング数: {len(image_map)}件')

# 年度別集計
year_counts = {}
for qid in image_map:
    y = qid[:4]
    year_counts[y] = year_counts.get(y, 0) + 1

for y, c in sorted(year_counts.items()):
    print(f'  {y}年: {c}枚')
